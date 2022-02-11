from os import path, mkdir, remove 
from pathlib import Path
from datetime import date, datetime
from orden import Orden
import time
import openpyxl as op

src_path = Path(__file__).parent
dowload_path = src_path.parent / 'download'
excel_path = src_path.parent / 'excel'
bot_path = src_path.parent / 'invoice bot'


# Verifica si existe el archivo excel 
def existe_archivo_txt(nombre_archivo):
    if path.exists(f'{src_path.parent}\\{nombre_archivo}.txt'):
        return True
    else:
        return False

def existe_archivo_excel(nombre_archivo):
    if path.exists(f'{excel_path}\\{nombre_archivo}.xlsx'):
        return True
    else: 
        raise Exception(f"Archivo no encontrado en la carpeta {excel_path}") 

def verificacion_carpetas():
    if path.exists(dowload_path) and path.exists(excel_path) and path.exists(bot_path):
        True
    else:
        print("Creacion de carpetas esenciales")
        if not path.exists(dowload_path):
            mkdir(f'{dowload_path}')
        if not path.exists(excel_path):
            mkdir(f'{excel_path}')
        print("Carpetas creadas")

def crear_carpeta():
    
    secuencia = 1
    condicion = False

    if not path.exists(f'{dowload_path}\\invoices-{date.today()}'):
        print(f'Se creo la carpeta .\invoices-{date.today()} en la carpeta .\download')
        mkdir(f'{dowload_path}\\invoices-{date.today()}')
        ruta = f'{dowload_path}\\invoices-{date.today()}'
        return ruta
    else: 
        while (condicion == False):
            if not path.exists(f'{dowload_path}\\invoices-{date.today()}-{secuencia}'):
                print(f'Se creo la carpeta .\invoices-{date.today()}-{secuencia} en la carpeta .\download.')
                mkdir(f'{dowload_path}\\invoices-{date.today()}-{secuencia}')
                nueva_ruta = f'{dowload_path}\\invoices-{date.today()}-{secuencia}'
                condicion = True
                return nueva_ruta
            else: 
                secuencia += 1
    
def escribir_texto(orden):
    if path.exists(f'{src_path.parent}\\log.txt') == False:
        with open(f'{src_path.parent}\\log.txt', 'w') as error:
            error.write(f'*{datetime.today()} ---- Ha ocurrido un error con la orden {orden["nombre"]}' + "\n")
    else: 
        with open(f'{src_path.parent}\\log.txt', 'a') as error:
            error.write(f'*{datetime.today()} ---- Ha ocurrido un error con la orden {orden["nombre"]}' + "\n")

def eliminar_archivo_texto():
    if path.exists(f'{src_path.parent}\\log.txt'):
        remove(f'{src_path.parent}\\log.txt')

def obtencion_columnas(archivo_excel):
    
    cabeceras = ["orden","user","password"] 
    nr_columnas = {}
    
    for cabecera in cabeceras:
        for columna in range(1, archivo_excel.max_column + 1):
            if cabecera == archivo_excel.cell(row = 1, column = columna).value :
                nr_columnas[cabecera]=columna

    return nr_columnas


def lectura_lista_orden(nombre_archivo):
    
    # Abro un archivo leo su hoja principal
    archivo_lista_invoice = op.load_workbook(f'{excel_path}\\{nombre_archivo}.xlsx')
    lista_invoice = archivo_lista_invoice.active

    columnas = obtencion_columnas(lista_invoice)

    invoices = []

    for orden in range(2, lista_invoice.max_row + 1):
        try:
            invoices.append(Orden(lista_invoice.cell(row=orden, column= columnas["user"]).value,
                                  lista_invoice.cell(row=orden, column = columnas["password"]).value,
                                 {"nombre":str(lista_invoice.cell(row = orden, column= columnas["orden"]).value).replace("|","").rstrip(),
                                  "link": lista_invoice.cell(row = orden, column= columnas["orden"]).hyperlink.target}))

        except Exception as ex:
            pass
    
    archivo_lista_invoice.close()

    return invoices

# Serials

def guardar_archivo_excel(archivo, nombre_archivo):
    # Ultimo paso coloco cabeceras a las columnas, y modifica tamaño de celdas
    informacion = archivo.get_sheet_by_name('Informacion')
    informacion.cell(row = 1, column = 1).value = "Nº ORDEN"
    informacion.cell(row = 1, column = 2).value = "USUARIO"
    informacion.cell(row = 1, column = 3).value = "ORDER DATE"
    informacion.cell(row = 1, column = 4).value = "INVOICE DATE"
    informacion.cell(row = 1, column = 5).value = "TARJETA"
    informacion.cell(row = 1, column = 6).value = "Nº TARJETA"
    informacion.cell(row = 1, column = 7).value = "TOTAL"

    hoja = archivo.get_sheet_by_name('Serials')

    hoja.cell(row = 1, column = 1).value = "Nº ORDEN"
    hoja.cell(row = 1, column = 2).value = "MODELO"

    numero_serial = 1

    for columna in range(3, hoja.max_column + 1):     
        hoja.cell(row = 1, column = columna).value = f'SERIAL {numero_serial}'
        hoja.column_dimensions[op.utils.get_column_letter(columna)].width = 20
        numero_serial+=1
    
    errores = archivo.get_sheet_by_name("Errores")

    errores.cell(row = 1, column = 1).value = "FECHA"
    errores.cell(row = 1, column = 2).value = "HORA"
    errores.cell(row = 1, column = 3).value = "Nº ORDEN"
    errores.cell(row = 1, column = 3).value = "MENSAJE"

    archivo.save(f'{excel_path}\\{date.today()}-{nombre_archivo}.xlsx')



def informacion_repetida(archivo_a_leer, hoja, primer_dato_lista):
    # Solo verifica la informacion del primer serial para ver si se repite
    hoja_a_leer = archivo_a_leer.get_sheet_by_name(hoja)

    for fila in range(2, hoja_a_leer.max_row + 1):
        if hoja_a_leer.cell(row = fila, column = 4).value == primer_dato_lista:
            return False
 
    return True


def crear_archivo(nombre_archivo):
    if path.exists(f'{excel_path}\\{date.today()}-{nombre_archivo}.xlsx'):
        archivo = op.load_workbook(f'{excel_path}\\{date.today()}-{nombre_archivo}.xlsx')
        return archivo
    else: 
        # Se crea un nuevo archivo 
        archivo = op.Workbook()
        # Se renombra la primera hoja
        archivo.worksheets[0].title = "Informacion"
        # Me muevo a la hoja
        informacion = archivo.get_sheet_by_name("Informacion")
        # Tamaños predeterminados dados
        informacion.column_dimensions['A'].width = 30
        informacion.column_dimensions['B'].width = 40
        informacion.column_dimensions['C'].width = 40
        informacion.column_dimensions['D'].width = 35
        informacion.column_dimensions['E'].width = 35
        informacion.column_dimensions['F'].width = 35
        informacion.column_dimensions['G'].width = 20
        # Creo nueva hoja 
        archivo.create_sheet('Serials')
        # Me muevo a la nueva hoja
        serial = archivo.get_sheet_by_name("Serials")
        serial.column_dimensions['A'].width = 30
        serial.column_dimensions['B'].width = 40
        serial.column_dimensions['C'].width = 40
        # Creo nueva hoja
        archivo.create_sheet('Errores')
        errores = archivo.get_sheet_by_name("Errores")
        errores.column_dimensions['A'].width = 20
        errores.column_dimensions['B'].width = 20
        errores.column_dimensions['C'].width = 20
        errores.column_dimensions['D'].width = 35
        
        return archivo

def escribir_serial(archivo_a_modificar, hoja, orden, nombre_producto, serials):

    if informacion_repetida(archivo_a_modificar, hoja, serials[0]):
        #Busco la hoja donde tengo que modificar 
        hoja_a_modificar = archivo_a_modificar.get_sheet_by_name(hoja)
        #Selecciono la ultima fila para asi no tener que guardar en un lugar especifico
        ultima_fila = hoja_a_modificar.max_row
        #Escribo en la columna uno que es el numero de orden:
        hoja_a_modificar.cell(row = ultima_fila+1, column = 1).hyperlink = orden["link"] #Hay que ver como guardar el orden
        hoja_a_modificar.cell(row = ultima_fila+1, column = 1).value = orden["nombre"]
        hoja_a_modificar.cell(row = ultima_fila+1, column = 1).style = 'Hyperlink'
        #hoja_a_modificar.cell(row = ultima_fila+1, column = 1).font = Font(underline='single')
        #Escribo en la columna tres que es el nombre del comprado:
        hoja_a_modificar.cell(row = ultima_fila+1, column = 2).value = nombre_producto
        
        #En las columnas siguientes se escribira los serials:
        columna_comienzo = 3
        
        for serial in serials:
            hoja_a_modificar.cell(row = ultima_fila+1, column = columna_comienzo).value = serial
            columna_comienzo +=1

def escribir_informacion(archivo_a_modificar, hoja, orden, user, order_date, invoice_date, credit_card, number_card, total):

    hoja_a_modificar = archivo_a_modificar.get_sheet_by_name(hoja)
    #Selecciono la ultima fila para asi no tener que guardar en un lugar especifico
    ultima_fila = hoja_a_modificar.max_row

    # Relleno la informacion necesaria
    # Orden 
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).hyperlink = orden["link"] #Hay que ver como guardar el orden
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).value = orden["nombre"]
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).style = 'Hyperlink'
    # User
    hoja_a_modificar.cell(row = ultima_fila+1, column = 2).value = user
    # Order date 
    hoja_a_modificar.cell(row = ultima_fila+1, column = 3).value = order_date
    # Invoice date
    hoja_a_modificar.cell(row = ultima_fila+1, column = 4).value = invoice_date
    # Credit card
    hoja_a_modificar.cell(row = ultima_fila+1, column = 5).value = credit_card
    # Numero de tarjeta 
    hoja_a_modificar.cell(row = ultima_fila+1, column = 6).value = number_card
    # Total 
    hoja_a_modificar.cell(row = ultima_fila+1, column = 7).value = total

def escribir_errores(archivo_a_modificar, hoja, error):
    # Relleno las filas 
    hoja_a_modificar = archivo_a_modificar.get_sheet_by_name(hoja)

    ultima_fila = hoja_a_modificar.max_row
    
    hoja_a_modificar.cell(row = ultima_fila+1, column = 1).value = date.today().strftime('%Y-%m-%d')
    hoja_a_modificar.cell(row = ultima_fila+1, column = 2).value = time.strftime('%H:%M:%S', time.localtime())
    hoja_a_modificar.cell(row = ultima_fila+1, column = 3).value = error["nombre"]
    hoja_a_modificar.cell(row = ultima_fila+1, column = 4).value = "No se ha podido desargar este invoice"